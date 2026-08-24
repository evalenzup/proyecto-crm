# app/services/conciliacion_service.py
"""Conciliación bancaria: importar, cotejar y exportar.

Sustituye el Excel que se llenaba a mano. La lógica de negocio es deliberadamente
poca: el sistema no adivina qué factura corresponde a qué depósito. Quien
concilia es la persona; el sistema aporta el orden, la búsqueda por folio, la
suma de control y que nada se pierda entre sesiones.
"""
from __future__ import annotations

import logging
import os
import uuid
from datetime import date
from decimal import Decimal
from io import BytesIO
from typing import List, Optional
from uuid import UUID

from fastapi import HTTPException
from sqlalchemy import func, or_
from sqlalchemy.orm import Session, selectinload

from app.config import settings
from app.models.conciliacion import (
    ConciliacionBancaria, MovimientoBancario, MovimientoEgreso, MovimientoFactura,
)
from app.models.cliente import Cliente
from app.models.empresa import Empresa
from app.models.egreso import Egreso
from app.models.factura import Factura
from app.services import estado_cuenta_banamex as lector

logger = logging.getLogger("app")

DIR_ESTADOS = os.path.join(settings.DATA_DIR, "estados_cuenta")


# ── Empresas que comparten cuenta ────────────────────────────────────────────

def _empresas_hermanas(db: Session, empresa_id: UUID) -> List[UUID]:
    """Empresas que comparten RFC con la dada.

    Fumigaciones, Jardinería y Limpieza facturan con el mismo RFC y cobran en la
    misma cuenta, así que un depósito puede corresponder a una factura de
    cualquiera de las tres. Buscar sólo en la empresa activa dejaría fuera dos
    terceras partes de las facturas.
    """
    rfc = db.query(Empresa.rfc).filter(Empresa.id == empresa_id).scalar()
    if not rfc:
        return [empresa_id]
    return [r[0] for r in db.query(Empresa.id).filter(Empresa.rfc == rfc).all()]


# ── Importar ─────────────────────────────────────────────────────────────────

def importar(
    db: Session,
    *,
    empresa_id: UUID,
    pdf_bytes: bytes,
    nombre_archivo: str,
    usuario_id: Optional[UUID] = None,
) -> ConciliacionBancaria:
    """Lee el PDF, valida que cuadre y lo archiva junto con sus movimientos."""
    try:
        estado = lector.leer(pdf_bytes)
    except lector.EstadoCuentaInvalido as exc:
        raise HTTPException(status_code=422, detail=str(exc))

    ya = (
        db.query(ConciliacionBancaria)
        .filter(
            ConciliacionBancaria.empresa_id == empresa_id,
            ConciliacionBancaria.cuenta == estado.cuenta,
            ConciliacionBancaria.periodo_inicio == estado.periodo_inicio,
            ConciliacionBancaria.periodo_fin == estado.periodo_fin,
        )
        .first()
    )
    if ya:
        raise HTTPException(
            status_code=409,
            detail=(
                f"Ya existe la conciliación del {estado.periodo_inicio:%d/%m/%Y} al "
                f"{estado.periodo_fin:%d/%m/%Y} para esta cuenta. Ábrela en vez de "
                "volver a importarla; si necesitas empezar de nuevo, elimínala primero."
            ),
        )

    conc = ConciliacionBancaria(
        empresa_id=empresa_id,
        periodo_inicio=estado.periodo_inicio,
        periodo_fin=estado.periodo_fin,
        banco="BANAMEX",
        cuenta=estado.cuenta,
        archivo_nombre=nombre_archivo,
        saldo_inicial=estado.saldo_inicial,
        saldo_final=estado.saldo_final,
        total_depositos=estado.total_depositos,
        total_retiros=estado.total_retiros,
        n_depositos=estado.n_depositos,
        n_retiros=estado.n_retiros,
        creado_por=usuario_id,
    )
    db.add(conc)
    db.flush()

    # El PDF se conserva: es el respaldo del trabajo ante la contadora
    os.makedirs(DIR_ESTADOS, exist_ok=True)
    destino = os.path.join(DIR_ESTADOS, f"{conc.id}.pdf")
    with open(destino, "wb") as f:
        f.write(pdf_bytes)
    conc.archivo_path = destino

    for i, m in enumerate(estado.movimientos, start=1):
        db.add(MovimientoBancario(
            conciliacion_id=conc.id, orden=i, fecha=m.fecha,
            concepto=m.concepto, deposito=m.deposito, retiro=m.retiro,
        ))

    db.commit()
    db.refresh(conc)
    logger.info("[Conciliación] importado %s–%s: %d movimientos",
                estado.periodo_inicio, estado.periodo_fin, len(estado.movimientos))
    return conc


def obtener(db: Session, conciliacion_id: UUID) -> ConciliacionBancaria:
    conc = (
        db.query(ConciliacionBancaria)
        .options(selectinload(ConciliacionBancaria.movimientos)
                 .selectinload(MovimientoBancario.facturas),
                 selectinload(ConciliacionBancaria.movimientos)
                 .selectinload(MovimientoBancario.egresos))
        .filter(ConciliacionBancaria.id == conciliacion_id)
        .first()
    )
    if not conc:
        raise HTTPException(status_code=404, detail="Conciliación no encontrada")
    return conc


def listar(db: Session, empresa_id: UUID) -> List[ConciliacionBancaria]:
    return (
        db.query(ConciliacionBancaria)
        .options(selectinload(ConciliacionBancaria.movimientos))
        .filter(ConciliacionBancaria.empresa_id.in_(_empresas_hermanas(db, empresa_id)))
        .order_by(ConciliacionBancaria.periodo_inicio.desc())
        .all()
    )


def eliminar(db: Session, conciliacion_id: UUID) -> None:
    conc = obtener(db, conciliacion_id)
    if conc.archivo_path and os.path.exists(conc.archivo_path):
        try:
            os.remove(conc.archivo_path)
        except OSError:
            logger.warning("No se pudo borrar %s", conc.archivo_path)
    db.delete(conc)
    db.commit()


# ── Trabajo sobre cada movimiento ────────────────────────────────────────────

def obtener_movimiento(db: Session, movimiento_id: UUID) -> MovimientoBancario:
    mov = db.query(MovimientoBancario).filter(
        MovimientoBancario.id == movimiento_id).first()
    if not mov:
        raise HTTPException(status_code=404, detail="Movimiento no encontrado")
    return mov


def actualizar_movimiento(db: Session, movimiento_id: UUID, datos: dict) -> MovimientoBancario:
    mov = obtener_movimiento(db, movimiento_id)
    for campo in ("comentario", "area", "conciliado"):
        if campo in datos:
            setattr(mov, campo, datos[campo])
    db.commit()
    db.refresh(mov)
    return mov


def enlazar_facturas(db: Session, movimiento_id: UUID, factura_ids: List[UUID]) -> MovimientoBancario:
    """Fija qué facturas componen el movimiento y arma el comentario.

    El comentario se rehace con los folios porque es lo que la contadora lee.
    Si la persona ya había escrito una nota a mano y no hay facturas, no se
    toca: ahí el texto es el dato.
    """
    mov = obtener_movimiento(db, movimiento_id)

    db.query(MovimientoFactura).filter(
        MovimientoFactura.movimiento_id == mov.id).delete(synchronize_session=False)
    db.flush()

    facturas = []
    if factura_ids:
        facturas = db.query(Factura).filter(Factura.id.in_(factura_ids)).all()
        faltan = set(factura_ids) - {f.id for f in facturas}
        if faltan:
            raise HTTPException(status_code=404, detail="Alguna factura ya no existe.")
        for f in facturas:
            db.add(MovimientoFactura(movimiento_id=mov.id, factura_id=f.id))

    if facturas:
        orden = sorted(facturas, key=lambda f: (f.serie or "", f.folio or 0))
        mov.comentario = ", ".join(f"{f.serie}-{f.folio}" for f in orden)
        mov.conciliado = True
    elif mov.conciliado and not (mov.comentario or "").strip():
        mov.conciliado = False

    db.commit()
    db.refresh(mov)
    return mov


def enlazar_egresos(db: Session, movimiento_id: UUID, egreso_ids: List[UUID]) -> MovimientoBancario:
    """Fija qué gastos componen el retiro y arma el comentario con su descripción."""
    mov = obtener_movimiento(db, movimiento_id)

    db.query(MovimientoEgreso).filter(
        MovimientoEgreso.movimiento_id == mov.id).delete(synchronize_session=False)
    db.flush()

    egresos = []
    if egreso_ids:
        egresos = db.query(Egreso).filter(Egreso.id.in_(egreso_ids)).all()
        if set(egreso_ids) - {e.id for e in egresos}:
            raise HTTPException(status_code=404, detail="Algún egreso ya no existe.")
        for e in egresos:
            db.add(MovimientoEgreso(movimiento_id=mov.id, egreso_id=e.id))

    if egresos:
        partes = []
        for e in egresos:
            etiqueta = (e.proveedor or "").strip() or (e.descripcion or "").strip()
            partes.append(etiqueta[:60])
        mov.comentario = " · ".join(partes)
        mov.conciliado = True
    elif mov.conciliado and not (mov.comentario or "").strip():
        mov.conciliado = False

    db.commit()
    db.refresh(mov)
    return mov


def buscar_egresos(db: Session, *, empresa_id: UUID, q: str = "", limite: int = 25) -> List[Egreso]:
    """Busca gastos para enlazar a un retiro, por proveedor o descripción."""
    empresas = _empresas_hermanas(db, empresa_id)
    consulta = (
        db.query(Egreso)
        .options(selectinload(Egreso.empresa))
        .filter(Egreso.empresa_id.in_(empresas))
    )
    q = (q or "").strip()
    if q:
        patron = f"%{q.lower()}%"
        consulta = consulta.filter(or_(
            func.lower(func.coalesce(Egreso.proveedor, "")).like(patron),
            func.lower(func.coalesce(Egreso.descripcion, "")).like(patron),
        ))
    return consulta.order_by(Egreso.fecha_egreso.desc()).limit(limite).all()


def buscar_facturas(
    db: Session, *, empresa_id: UUID, q: str = "", limite: int = 25,
) -> List[Factura]:
    """Busca facturas para enlazar. Por folio primero, que es como se piden.

    Acepta "1585", "A-1585" o "a1585", y si no parece folio busca por cliente.
    """
    empresas = _empresas_hermanas(db, empresa_id)
    consulta = (
        db.query(Factura)
        .options(selectinload(Factura.cliente), selectinload(Factura.empresa))
        .filter(Factura.empresa_id.in_(empresas), Factura.estatus != "BORRADOR")
    )

    q = (q or "").strip()
    if q:
        solo_digitos = "".join(c for c in q if c.isdigit())
        condiciones = []
        if solo_digitos:
            condiciones.append(Factura.folio == int(solo_digitos))
        condiciones.append(Factura.cliente.has(
            func.lower(func.coalesce(Cliente.nombre_comercial, "")).contains(q.lower())
        ))
        consulta = consulta.filter(or_(*condiciones))

    return consulta.order_by(Factura.folio.desc()).limit(limite).all()


# ── Exportar ─────────────────────────────────────────────────────────────────

def exportar_excel(db: Session, conciliacion_id: UUID) -> BytesIO:
    """Arma el archivo que se le manda a la contadora.

    Una sola hoja con las mismas columnas que ella ya recibe, para que del otro
    lado no cambie nada.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    conc = obtener(db, conciliacion_id)
    wb = Workbook()
    ws = wb.active
    ws.title = "CONCILIACIÓN"

    negrita = Font(bold=True)
    titulo = Font(bold=True, size=12)
    encabezado = PatternFill("solid", fgColor="D9D9D9")
    borde = Border(*[Side(style="thin", color="999999")] * 4)
    money = '#,##0.00'

    ws["A1"] = "Movimiento de cuenta Cheques"
    ws["A1"].font = titulo
    ws["A3"] = "Resumen de cuenta"; ws["A3"].font = negrita
    ws["A4"], ws["B4"] = "Cuenta", conc.cuenta or ""
    ws["D4"], ws["E4"] = "Empresa", conc.empresa.nombre_comercial if conc.empresa else ""
    ws["A5"], ws["B5"] = "Banco", conc.banco
    ws["D5"], ws["E5"] = "Moneda", "MXN"
    ws["A7"] = (f"Resumen del {conc.periodo_inicio:%d/%m/%Y} "
                f"al {conc.periodo_fin:%d/%m/%Y}")
    ws["A7"].font = negrita
    ws["A8"], ws["B8"] = "Saldo inicial", float(conc.saldo_inicial)
    ws["D8"], ws["E8"] = "Saldo final", float(conc.saldo_final)
    ws["A9"], ws["B9"] = f"Depósitos ({conc.n_depositos})", float(conc.total_depositos)
    ws["A10"], ws["B10"] = f"Retiros ({conc.n_retiros})", float(conc.total_retiros)
    for celda in ("B8", "E8", "B9", "B10"):
        ws[celda].number_format = money

    ws["A12"] = "Detalle de Movimientos"
    ws["A12"].font = negrita

    columnas = ["Fecha", "Descripción", "Depósitos", "Retiros",
                "Comentarios", "Área", "Facturas"]
    for i, nombre in enumerate(columnas, start=1):
        c = ws.cell(row=13, column=i, value=nombre)
        c.font = negrita
        c.fill = encabezado
        c.border = borde

    fila = 14
    for m in conc.movimientos:
        ws.cell(row=fila, column=1, value=m.fecha).number_format = "yyyy-mm-dd"
        ws.cell(row=fila, column=2, value=m.concepto)
        ws.cell(row=fila, column=3,
                value=float(m.deposito) if m.deposito else None).number_format = money
        ws.cell(row=fila, column=4,
                value=float(m.retiro) if m.retiro else None).number_format = money
        ws.cell(row=fila, column=5, value=m.comentario or "")
        ws.cell(row=fila, column=6, value=m.area or "")
        # Los folios se escriben como texto ("A-1585") para que Excel no los
        # convierta en número: así llegaban como "1,585.00" y se leían mal.
        ws.cell(row=fila, column=7,
                value=", ".join(f"{f.serie}-{f.folio}" for f in m.facturas))
        for col in range(1, 8):
            ws.cell(row=fila, column=col).border = borde
            ws.cell(row=fila, column=col).alignment = Alignment(
                vertical="top", wrap_text=(col in (2, 5)))
        fila += 1

    for col, ancho in zip("ABCDEFG", (12, 58, 14, 14, 34, 10, 26)):
        ws.column_dimensions[col].width = ancho
    ws.freeze_panes = "A14"

    salida = BytesIO()
    wb.save(salida)
    salida.seek(0)
    return salida
