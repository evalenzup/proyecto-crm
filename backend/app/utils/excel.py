from io import BytesIO
from typing import List, Dict, Any
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def generate_excel(data: List[Dict[str, Any]], headers: Dict[str, str], sheet_name: str = "Datos") -> BytesIO:
    """
    Genera un archivo Excel en memoria a partir de una lista de diccionarios.

    :param data: Lista de diccionarios con los datos.
    :param headers: Diccionario donde key = llave del dato y value = título de columna.
                    El orden de este diccionario dicta el orden de las columnas.
    :param sheet_name: Nombre de la hoja de cálculo.
    :return: BytesIO con el contenido del archivo Excel.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    # Definir estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    thin_border = Side(border_style="thin", color="000000")
    border = Border(left=thin_border, right=thin_border, top=thin_border, bottom=thin_border)

    # Convertir headers a lista de claves y títulos
    keys = list(headers.keys())
    titles = list(headers.values())

    # Escribir encabezados
    for col_num, title in enumerate(titles, 1):
        cell = ws.cell(row=1, column=col_num, value=title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border

    # Escribir datos
    for row_num, item in enumerate(data, 2):
        for col_num, key in enumerate(keys, 1):
            val = item.get(key)
            # Manejo básico de tipos (None -> '')
            if val is None:
                val = ""
            # Convertir booleanos a Sí/No si se desea, o dejarlos como str
            if isinstance(val, bool):
                val = "Sí" if val else "No"
            
            cell = ws.cell(row=row_num, column=col_num, value=str(val) if not isinstance(val, (int, float)) else val)
            cell.border = border
            # Alinear a la izquierda por defecto, números a la derecha se hace auto en Excel a veces, 
            # pero openpyxl no lo fuerza a menos que lo especifiques. Dejamos default.

    # Ajustar ancho de columnas (automático simple)
    for col_num, key in enumerate(keys, 1):
        # Calculamos longitud máxima basada en header y primeros 50 datos para no tardar mucho
        max_length = len(titles[col_num-1])
        for i, item in enumerate(data[:50]):
            val = item.get(key)
            if val:
                max_length = max(max_length, len(str(val)))
        
        adjusted_width = (max_length + 2)
        column_letter = ws.cell(row=1, column=col_num).column_letter
        ws.column_dimensions[column_letter].width = min(adjusted_width, 50) # limite 50 chars ancho

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output
